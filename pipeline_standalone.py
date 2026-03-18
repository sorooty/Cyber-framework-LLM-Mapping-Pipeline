"""
pipeline_standalone.py — Pipeline de mapping référentiels en fichier unique.

Consolide : models, config, survey_validator, schemas, adaptateur générique
unifié (plus d'adaptateur CIS/NIST spécifique), ingester, similarity (avec
cache du modèle d'embedding au niveau module), heatmap, scorer LLM en mode
batch (N paires par appel API), pipeline.

Adaptateur unifié :
    load_generic_yaml gère tous les référentiels (CIS, NIST, ISO…).
    section_id_template dans FrameworkSchema applique le préfixe (ex. "CIS-1").

LLM batching :
    LLM_BATCH_SIZE paires regroupées par appel → ~200 appels → ~20 appels.

Cache modèle :
    _get_embedding_model() charge SentenceTransformer une seule fois par
    processus (cache module-level), évitant les rechargements redondants.

Usage :
    python pipeline_standalone.py \
        --ref-a files/surveys/cis-controls-v8-1.yaml --adapter-a generic \
        --ref-b files/surveys/nistCsfV2.yaml          --adapter-b generic

Options :
    --force-step1   Repaser les fichiers sources même si le cache existe
    --force-step2   Recalculer les embeddings même si le cache existe
    --force-step4   Rescorer avec le LLM même si le cache existe
    --skip-step4    Ne pas lancer le LLM (étapes 1-3 uniquement)
"""

# ══════════════════════════════════════════════════════════════════════════════
# Imports
# ══════════════════════════════════════════════════════════════════════════════

import argparse
import asyncio
import json
import os
import re
import sys
import yaml
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal, Optional, Union

import dotenv
dotenv.load_dotenv()

from openai import AsyncOpenAI
from pydantic import BaseModel, Field, ValidationError
from sentence_transformers import SentenceTransformer


# ══════════════════════════════════════════════════════════════════════════════
# Config
# ══════════════════════════════════════════════════════════════════════════════

DATA_DIR    = Path(__file__).parent / "data"
OUTPUT_DIR  = DATA_DIR / "outputs"
LOGS_DIR    = DATA_DIR / "logs"
SURVEYS_DIR = Path(__file__).parent / "files" / "surveys"
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)

SURVEY_TS_PATH = Path(__file__).parent / "files" / "survey.ts"

# Étape 2 : Similarité sémantique
EMBEDDING_MODEL    = "paraphrase-multilingual-MiniLM-L12-v2"  # multilingue FR/EN
SEMANTIC_THRESHOLD = 0.40
TOP_K              = 5

# Étape 4 : LLM
LLM_MODEL             = "gpt-4o-mini"
LLM_MAX_RETRIES       = 2
LLM_CONFIRM_THRESHOLD = 0.75
LLM_CONCURRENCY       = 20   # gpt-4o-mini tolère facilement 20 req. parallèles
LLM_BATCH_SIZE        = 10   # paires par appel API (1 appel → N paires)

# Étape 5 : Nettoyage
LLM_MIN_CONFIDENCE = 0.40    # confiance minimale pour garder une relation

# Seuils relation_type (centralisés ici, utilisés par _infer_relation_type)
THRESHOLD_EQUIVALENCE = 0.85
THRESHOLD_COVERAGE    = 0.75


# ══════════════════════════════════════════════════════════════════════════════
# Models
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class RequirementNormalized:
    id: str
    framework: str
    title: str
    description: str
    tags: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "id":          self.id,
            "framework":   self.framework,
            "title":       self.title,
            "description": self.description,
            "tags":        self.tags,
        }

    @staticmethod
    def from_dict(d: dict) -> "RequirementNormalized":
        return RequirementNormalized(
            id=d["id"],
            framework=d["framework"],
            title=d["title"],
            description=d.get("description", ""),
            tags=d.get("tags", []),
        )


@dataclass
class CandidatePair:
    id_A: str
    title_A: str
    id_B: str
    title_B: str
    semantic_score: float

    def to_dict(self) -> dict:
        return {
            "id_A":           self.id_A,
            "title_A":        self.title_A,
            "id_B":           self.id_B,
            "title_B":        self.title_B,
            "semantic_score": round(self.semantic_score, 4),
        }


@dataclass
class MappingRelation:
    id_A: str
    title_A: str
    id_B: str
    title_B: str
    semantic_score: float
    coverage_A_to_B: float
    coverage_B_to_A: float
    confidence: float
    relation_type: str   # "equivalence"|"A_couvre_B"|"B_couvre_A"|"partielle"|"aucun_lien"
    justification: str = ""
    framework_A: str = ""   # clé canonique du framework source (ex. "CIS_v8")
    framework_B: str = ""   # clé canonique du framework cible (ex. "ISO_27001")

    def to_dict(self) -> dict:
        return {
            "framework_A":     self.framework_A,
            "framework_B":     self.framework_B,
            "id_A":            self.id_A,
            "title_A":         self.title_A,
            "id_B":            self.id_B,
            "title_B":         self.title_B,
            "semantic_score":  round(self.semantic_score, 4),
            "coverage_A_to_B": round(self.coverage_A_to_B, 4),
            "coverage_B_to_A": round(self.coverage_B_to_A, 4),
            "confidence":      round(self.confidence, 4),
            "relation_type":   self.relation_type,
            "justification":   self.justification,
        }


# ── Helpers de classification (réutilisés partout) ────────────────────────────

def _infer_relation_type(cov_a: float, cov_b: float) -> str:
    """Dérive relation_type depuis les scores de coverage."""
    if cov_a >= THRESHOLD_EQUIVALENCE and cov_b >= THRESHOLD_EQUIVALENCE:
        return "equivalence"
    if cov_a >= THRESHOLD_COVERAGE and cov_b < THRESHOLD_COVERAGE:
        return "A_couvre_B"
    if cov_b >= THRESHOLD_COVERAGE and cov_a < THRESHOLD_COVERAGE:
        return "B_couvre_A"
    if cov_a >= 0.4 or cov_b >= 0.4:
        return "partielle"
    return "aucun_lien"



def _parse_survey_ts(path: Path = SURVEY_TS_PATH) -> dict:
    src = path.read_text(encoding="utf-8")
    m = re.search(r"defineSurveyConfig\(\{(.+?)\}\)\s*\ntype", src, re.DOTALL)
    if not m:
        raise ValueError("Impossible de localiser defineSurveyConfig({...}) dans survey.ts")
    raw = "{" + m.group(1) + "}"
    raw = re.sub(r"//[^\n]*", "", raw)
    raw = raw.replace("'", '"')
    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    raw = re.sub(r'(?<!")(\b[a-zA-Z_][a-zA-Z0-9_]*\b)(?!")\s*:', r'"\1":', raw)
    raw = re.sub(r'(?<!")(\b\d+\b)(?!")\s*:', r'"\1":', raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        lines = raw.splitlines()
        err_line = e.lineno - 1
        context = "\n".join(lines[max(0, err_line - 3): err_line + 3])
        raise ValueError(f"Erreur de parsing JSON survey.ts (ligne {e.lineno}) :\n{context}\n\n{e}")


def _count_from_groups(groups: dict) -> dict[str, int]:
    counts: dict[str, int] = {}
    for sec, val in groups.items():
        if isinstance(val, int):
            counts[str(sec)] = val
        elif isinstance(val, dict):
            for sub, n in val.items():
                if isinstance(n, int):
                    counts[f"{sec}.{sub}"] = n
    return counts


def expected_counts(framework_key: str, survey: dict | None = None) -> dict[str, int]:
    if survey is None:
        survey = _parse_survey_ts()
    if framework_key not in survey:
        raise KeyError(
            f"'{framework_key}' introuvable dans survey.ts.\n"
            f"Clés disponibles : {sorted(survey.keys())}"
        )
    return _count_from_groups(survey[framework_key]["groups"])


def total_expected(framework_key: str, survey: dict | None = None) -> int:
    return sum(expected_counts(framework_key, survey).values())


@dataclass
class ValidationReport:
    framework_key: str
    total_expected: int
    total_parsed: int
    missing_sections: list[str] = field(default_factory=list)
    over_sections: list[str]    = field(default_factory=list)
    section_detail: dict        = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return self.total_expected == self.total_parsed

    def __str__(self) -> str:
        status = "✅ OK" if self.ok else "❌ ÉCART"
        lines = [
            f"\n{'='*60}",
            f" VALIDATION — {self.framework_key}  [{status}]",
            f"{'='*60}",
            f" Attendu  : {self.total_expected} exigences",
            f" Parsé    : {self.total_parsed} exigences",
            f" Écart    : {self.total_parsed - self.total_expected:+d}",
        ]
        if self.missing_sections:
            lines.append(f"\n Sections manquantes ({len(self.missing_sections)}) :")
            for s in self.missing_sections[:20]:
                exp = self.section_detail.get(s, (0, 0))
                lines.append(f"   {s:<15} attendu={exp[0]}  parsé={exp[1]}")
        if self.over_sections:
            lines.append(f"\n Sections en excès ({len(self.over_sections)}) :")
            for s in self.over_sections[:20]:
                exp = self.section_detail.get(s, (0, 0))
                lines.append(f"   {s:<15} attendu={exp[0]}  parsé={exp[1]}")
        lines.append("=" * 60)
        return "\n".join(lines)


def validate_survey(
    framework_key: str,
    parsed_requirements: list,
    survey: dict | None = None,
) -> ValidationReport:
    if survey is None:
        survey = _parse_survey_ts()
    exp = expected_counts(framework_key, survey)
    total_exp = sum(exp.values())
    parsed_by_section: dict[str, int] = {}
    for req in parsed_requirements:
        parts = req.id.split(".")
        if len(parts) >= 3:
            key2 = f"{parts[0]}.{parts[1]}"
            key1 = parts[0]
            if key2 in exp:
                parsed_by_section[key2] = parsed_by_section.get(key2, 0) + 1
            elif key1 in exp:
                parsed_by_section[key1] = parsed_by_section.get(key1, 0) + 1
            else:
                parsed_by_section[key2] = parsed_by_section.get(key2, 0) + 1
        elif len(parts) == 2:
            key = parts[0]
            parsed_by_section[key] = parsed_by_section.get(key, 0) + 1
        else:
            parsed_by_section[req.id] = parsed_by_section.get(req.id, 0) + 1
    section_detail = {
        s: (exp.get(s, 0), parsed_by_section.get(s, 0))
        for s in set(exp) | set(parsed_by_section)
    }
    missing = [s for s, (e, p) in section_detail.items() if e > 0 and p < e]
    over    = [s for s, (e, p) in section_detail.items() if p > e]
    return ValidationReport(
        framework_key=framework_key,
        total_expected=total_exp,
        total_parsed=len(parsed_requirements),
        missing_sections=sorted(missing),
        over_sections=sorted(over),
        section_detail=section_detail,
    )


YAML_TO_SURVEY_KEY: dict[str, str] = {
    "cis_controls_v8":     "cis-controls-v8-1",
    "nistCsfV2":           "nistCsfV2",
    "nis2":                "nis2",
    "nis2v2":              "nis2v2",
    "dora":                "dora",
    "iso27001":            "iso27001-2022",
    "iso9001":             "iso9001-2015",
    "iso13485":            "iso13485-2016",
    "iso14001":            "iso14001-2015",
    "iso17021":            "iso17021-2015",
    "iso17024":            "iso17024-2012",
    "iso20000":            "iso20000-2018",
    "iso20022_1":          "iso20022-1-2013",
    "iso20022_2":          "iso20022-2-2013",
    "iso26000":            "iso26000-2020",
    "iso27701":            "iso27701-2021",
    "iso42001":            "iso42001-2023",
    "iso50001":            "iso50001-2018",
    "hdsV2":               "hdsV2",
    "pcidssV4":            "pcidssV4",
    "rgpdCnil":            "rgpdCnil",
    "soc2_type2":          "soc2-type2",
    "qualiopiV9":          "qualiopiV9",
    "anssiHygieneGuideV2": "anssiHygieneGuideV2",
    "anssiAISecurity":     "anssiAISecurity",
    "owaspV4_0":           "owaspV4-0-3",
    "owasp_llm_ai_v1":     "owasp-llm-ai-v1",
    "eumdr":               "eumdr-2017",
    "ecc1":                "ecc1-2018",
    "5_20Law":             "5-20Law",
    "afnor_spec":          "afnor-spec-2217",
    "secnumcloud_v3":      "secnumcloud-v3-2",
    "igi11300":            "igi1300-2021",
    "ii901":               "ii901",
    "iec_62443_3_2":       "iec-62443-3-2-2020",
    "ichE6R2":             "ichE6R2",
    "isa6_0_3_tisax":      "isa6-0-3-tisax",
    "ai_act":              "ai-act",
    "fda_cfr_21_part":     "fda-cfr-21-part-111",
    "cosmetovigilance":    "cosmetovigilance-2009",
    "ansm":                "ansm-2022",
}


def validate_by_yaml_name(
    yaml_name: str,
    parsed_requirements: list,
    survey: dict | None = None,
) -> ValidationReport:
    if survey is None:
        survey = _parse_survey_ts()
    survey_key = YAML_TO_SURVEY_KEY.get(yaml_name)
    if survey_key is None:
        raise KeyError(
            f"Pas de correspondance survey.ts pour '{yaml_name}'.\n"
            f"Ajoute-le dans YAML_TO_SURVEY_KEY."
        )
    return validate_survey(survey_key, parsed_requirements, survey)


# ══════════════════════════════════════════════════════════════════════════════
# Schemas (FrameworkSchema)
# ══════════════════════════════════════════════════════════════════════════════

_FRAMEWORK_METADATA: dict[str, dict] = {
    "cis_controls_v8":     {"display_name": "CIS Controls v8",         "domain": "cybersecurity",  "language": "en", "section_id_prefix": "CIS-"},
    "nistCsfV2":           {"display_name": "NIST CSF v2",              "domain": "cybersecurity",  "language": "en"},
    "nis2":                {"display_name": "NIS2",                     "domain": "cybersecurity",  "language": "fr"},
    "nis2v2":              {"display_name": "NIS2 v2",                  "domain": "cybersecurity",  "language": "fr"},
    "dora":                {"display_name": "DORA",                     "domain": "financial",      "language": "fr"},
    "iso27001":            {"display_name": "ISO 27001:2022",           "domain": "cybersecurity",  "language": "en"},
    "iso27701":            {"display_name": "ISO 27701:2021",           "domain": "privacy",        "language": "en"},
    "iso9001":             {"display_name": "ISO 9001:2015",            "domain": "quality",        "language": "en"},
    "iso13485":            {"display_name": "ISO 13485:2016",           "domain": "medical",        "language": "en"},
    "iso14001":            {"display_name": "ISO 14001:2015",           "domain": "environment",    "language": "en"},
    "iso17021":            {"display_name": "ISO 17021:2015",           "domain": "audit",          "language": "en"},
    "iso17024":            {"display_name": "ISO 17024:2012",           "domain": "certification",  "language": "en"},
    "iso20000":            {"display_name": "ISO 20000:2018",           "domain": "itsm",           "language": "en"},
    "iso20022_1":          {"display_name": "ISO 20022-1:2013",         "domain": "financial",      "language": "en"},
    "iso20022_2":          {"display_name": "ISO 20022-2:2013",         "domain": "financial",      "language": "en"},
    "iso26000":            {"display_name": "ISO 26000:2020",           "domain": "social",         "language": "en"},
    "iso42001":            {"display_name": "ISO 42001:2023",           "domain": "ai",             "language": "en"},
    "iso50001":            {"display_name": "ISO 50001:2018",           "domain": "energy",         "language": "en"},
    "hdsV2":               {"display_name": "HDS v2",                   "domain": "healthcare",     "language": "fr"},
    "pcidssV4":            {"display_name": "PCI DSS v4",               "domain": "payment",        "language": "en"},
    "rgpdCnil":            {"display_name": "RGPD/CNIL",                "domain": "privacy",        "language": "fr"},
    "soc2_type2":          {"display_name": "SOC 2 Type II",            "domain": "cybersecurity",  "language": "en"},
    "qualiopiV9":          {"display_name": "Qualiopi v9",              "domain": "training",       "language": "fr"},
    "anssiHygieneGuideV2": {"display_name": "ANSSI Guide d'Hygiène v2", "domain": "cybersecurity",  "language": "fr"},
    "anssiAISecurity":     {"display_name": "ANSSI AI Security",        "domain": "ai",             "language": "fr"},
    "owaspV4_0":           {"display_name": "OWASP ASVS v4.0",          "domain": "cybersecurity",  "language": "en"},
    "owasp_llm_ai_v1":     {"display_name": "OWASP LLM AI v1",          "domain": "ai",             "language": "en"},
    "eumdr":               {"display_name": "EU MDR 2017",              "domain": "medical",        "language": "en"},
    "ecc1":                {"display_name": "ECC 1.0",                  "domain": "cybersecurity",  "language": "en"},
    "5_20Law":             {"display_name": "Loi 5-20",                 "domain": "law",            "language": "fr"},
    "afnor_spec":          {"display_name": "AFNOR Spec AI 22-17",      "domain": "ai",             "language": "fr"},
    "secnumcloud_v3":      {"display_name": "SecNumCloud v3",           "domain": "cloud",          "language": "fr"},
    "igi11300":            {"display_name": "IGI 1300:2021",            "domain": "classified",     "language": "fr"},
    "ii901":               {"display_name": "II 901",                   "domain": "cybersecurity",  "language": "fr"},
    "iec_62443_3_2":       {"display_name": "IEC 62443-3-2:2020",       "domain": "ot_security",    "language": "en"},
    "ichE6R2":             {"display_name": "ICH E6(R2)",               "domain": "medical",        "language": "en"},
    "isa6_0_3_tisax":      {"display_name": "ISA6 / TISAX",             "domain": "automotive",     "language": "en"},
    "ai_act":              {"display_name": "EU AI Act",                "domain": "ai",             "language": "mixed"},
    "fda_cfr_21_part":     {"display_name": "FDA CFR 21 Part 11",       "domain": "medical",        "language": "en"},
    "cosmetovigilance":    {"display_name": "Cosmétovigilance 2009",    "domain": "cosmetics",      "language": "fr"},
    "ansm":                {"display_name": "ANSM 2022",                "domain": "medical",        "language": "fr"},
}


@dataclass
class FrameworkSchema:
    key: str
    survey_key: str
    display_name: str
    groups: dict
    flat_groups: list[str]
    overwrite_ids: dict[int, int]
    language: str          = "en"
    domain: str            = "unknown"
    section_id_prefix: str = ""   # ex. "CIS-" → IDs de la forme "CIS-1.1"

    @property
    def total_expected(self) -> int:
        return sum(self._iter_counts())

    @property
    def all_sections(self) -> list[str]:
        result = []
        for sec, val in self.groups.items():
            if isinstance(val, int):
                result.append(str(sec))
            elif isinstance(val, dict):
                for sub in val:
                    result.append(f"{sec}.{sub}")
        return result

    def section_count(self, section: Union[int, str], subsection: Union[int, str, None] = None) -> int:
        sec = self.groups.get(section) or self.groups.get(str(section))
        if sec is None:
            return 0
        if isinstance(sec, int):
            return sec if subsection is None else 0
        if subsection is not None:
            return sec.get(subsection, 0) or sec.get(str(subsection), 0)
        return sum(n for n in sec.values() if isinstance(n, int))

    def effective_section_id(self, section_key: Union[int, str]) -> Union[int, str]:
        int_key = int(section_key) if str(section_key).isdigit() else section_key
        return self.overwrite_ids.get(int_key, int_key)

    def is_flat(self, section_key: Union[int, str]) -> bool:
        return str(section_key) in self.flat_groups

    def _iter_counts(self):
        for val in self.groups.values():
            if isinstance(val, int):
                yield val
            elif isinstance(val, dict):
                for n in val.values():
                    if isinstance(n, int):
                        yield n

    def __repr__(self) -> str:
        return f"FrameworkSchema(key={self.key!r}, total={self.total_expected}, domain={self.domain!r})"


def _build_schemas(survey: dict | None = None) -> dict[str, FrameworkSchema]:
    if survey is None:
        try:
            survey = _parse_survey_ts()
        except Exception:
            return {}
    schemas: dict[str, FrameworkSchema] = {}
    for yaml_key, survey_key in YAML_TO_SURVEY_KEY.items():
        if survey_key not in survey:
            continue
        cfg  = survey[survey_key]
        meta = _FRAMEWORK_METADATA.get(yaml_key, {})
        raw_overwrite = cfg.get("overwriteGroupsId", {})
        overwrite_ids = {int(k): int(v) for k, v in raw_overwrite.items()}
        schemas[yaml_key] = FrameworkSchema(
            key=yaml_key,
            survey_key=survey_key,
            display_name=meta.get("display_name", yaml_key),
            groups=cfg.get("groups", {}),
            flat_groups=cfg.get("flatGroups", []),
            overwrite_ids=overwrite_ids,
            language=meta.get("language", "en"),
            domain=meta.get("domain", "unknown"),
            section_id_prefix=meta.get("section_id_prefix", ""),
        )
    return schemas


SCHEMAS: dict[str, FrameworkSchema] = _build_schemas()


def get_schema(yaml_name: str) -> FrameworkSchema | None:
    return SCHEMAS.get(yaml_name)


# ══════════════════════════════════════════════════════════════════════════════
# Adapters
# ══════════════════════════════════════════════════════════════════════════════

# ── Generic YAML ──────────────────────────────────────────────────────────────

def _name_from_path(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"[-_]\d+$", "", stem)
    return stem.replace("-", "_").replace(" ", "_")


def _clean_id(raw: str) -> str:
    return re.sub(r"[^\x00-\x7F]", "", raw).strip(". ")


def _clean_text(raw: str) -> str:
    return re.sub(r"\s+", " ", raw).strip()


def _extract_generic(
    node: dict,
    framework: str,
    id_prefix: str,
    parent_tags: list[str],
    requirements: list[RequirementNormalized],
    schema: FrameworkSchema | None = None,
) -> None:
    all_nums = set()
    for k in node:
        if isinstance(k, str):
            m = k.split("_")
            if len(m) >= 2 and m[0].isdigit() and m[1] in ("label", "desc", "prefix"):
                all_nums.add(int(m[0]))

    for n in sorted(all_nums):
        label  = str(node.get(f"{n}_label",  "")).strip()
        desc   = str(node.get(f"{n}_desc",   "")).strip()
        prefix = str(node.get(f"{n}_prefix", "")).strip()

        if not label and desc:
            label = desc[:120].rstrip() + ("…" if len(desc) > 120 else "")
            desc  = ""

        if not label:
            continue

        req_id = prefix if prefix else f"{id_prefix}.{n}"
        req_id = _clean_id(req_id)

        requirements.append(RequirementNormalized(
            id=req_id,
            framework=framework,
            title=_clean_text(label),
            description=_clean_text(desc),
            tags=parent_tags.copy(),
        ))

    sub_nums = sorted(k for k in node if isinstance(k, int))
    for sub_num in sub_nums:
        sub = node[sub_num]
        if not isinstance(sub, dict):
            continue
        sub_title  = str(sub.get("title",  "")).strip()
        sub_prefix = str(sub.get("prefix", "")).strip()
        sub_id = sub_prefix if sub_prefix else f"{id_prefix}.{sub_num}"
        new_tags = parent_tags + ([sub_title] if sub_title else [])
        _extract_generic(
            node=sub,
            framework=framework,
            id_prefix=sub_id,
            parent_tags=new_tags,
            requirements=requirements,
            schema=schema,
        )


def load_generic_yaml(path: str, framework_name: str | None = None) -> list[RequirementNormalized]:
    path_obj = Path(path)
    framework = framework_name or _name_from_path(path_obj)
    schema    = get_schema(framework)

    with open(path_obj, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    requirements: list[RequirementNormalized] = []
    for section_num, section in data.items():
        if not isinstance(section, dict):
            continue
        section_title  = str(section.get("title",  "")).strip()
        section_prefix = str(section.get("prefix", "")).strip()
        effective_id   = str(schema.effective_section_id(section_num) if schema else section_num)
        if schema and schema.section_id_prefix:
            effective_id = f"{schema.section_id_prefix}{effective_id}"
        _extract_generic(
            node=section,
            framework=framework,
            id_prefix=section_prefix or effective_id,
            parent_tags=[section_title] if section_title else [],
            requirements=requirements,
            schema=schema,
        )
    return requirements


# Registre des adaptateurs nommés
ADAPTER_LOADERS: dict[str, callable] = {
    "generic": load_generic_yaml,
}


# ── Registry (découverte des référentiels disponibles) ────────────────────────

def list_frameworks() -> dict[str, Path]:
    """Retourne {display_name: yaml_path} pour tous les YAMLs disponibles."""
    frameworks = {}
    for yaml_path in sorted(SURVEYS_DIR.rglob("*.yaml")):
        if "__MACOSX" in yaml_path.parts:
            continue
        name = _name_from_path(yaml_path)
        frameworks[name] = yaml_path
    return frameworks


def load_framework(
    yaml_path: "str | Path",
    framework_name: "str | None" = None,
    validate: bool = True,
) -> list[RequirementNormalized]:
    """Charge un référentiel depuis son YAML via l'adaptateur générique."""
    requirements = load_generic_yaml(str(yaml_path), framework_name=framework_name)
    if validate and framework_name:
        try:
            report = validate_by_yaml_name(framework_name, requirements)
            if not report.ok:
                print(
                    f"  ⚠️  [survey.ts] {framework_name} : "
                    f"attendu={report.total_expected}, parsé={report.total_parsed} "
                    f"(écart {report.total_parsed - report.total_expected:+d})"
                )
        except (KeyError, Exception):
            pass
    return requirements


def get_expected_count(framework_name: str) -> "int | None":
    """Retourne le nombre d'exigences attendu selon survey.ts, ou None."""
    try:
        survey_key = YAML_TO_SURVEY_KEY.get(framework_name)
        if not survey_key:
            return None
        survey = _parse_survey_ts()
        return total_expected(survey_key, survey)
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# Étape 1 — Ingestion
# ══════════════════════════════════════════════════════════════════════════════

def _save_json(requirements: list[RequirementNormalized], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in requirements], f, ensure_ascii=False, indent=2)


def _load_json_requirements(path: Path) -> list[RequirementNormalized]:
    with open(path, "r", encoding="utf-8") as f:
        return [RequirementNormalized.from_dict(d) for d in json.load(f)]


def _validate_ingestion(source_stem: str, requirements: list[RequirementNormalized]) -> None:
    if not requirements:
        return
    framework_name = requirements[0].framework
    try:
        report = validate_by_yaml_name(framework_name, requirements)
        if report.ok:
            print(f"  [✅ survey.ts] {framework_name} : {report.total_parsed}/{report.total_expected}")
        else:
            print(
                f"  [⚠️  survey.ts] {framework_name} : "
                f"{report.total_parsed}/{report.total_expected} "
                f"(écart {report.total_parsed - report.total_expected:+d})"
            )
            for s in report.missing_sections[:3]:
                exp, got = report.section_detail[s]
                print(f"     section {s} manquante : attendu={exp}, parsé={got}")
    except KeyError:
        pass
    except Exception:
        pass


def _load_or_parse(
    source_path: str,
    loader,
    cache_path: Path,
    force: bool,
) -> list[RequirementNormalized]:
    if cache_path.exists() and not force:
        print(f"  [cache] {cache_path.name}")
        return _load_json_requirements(cache_path)
    print(f"  [parse] {Path(source_path).name} ...")
    requirements = loader(source_path)
    _save_json(requirements, cache_path)
    print(f"  [ok]    {len(requirements)} exigences -> {cache_path.name}")
    return requirements


def run_ingestion(
    path_A: str, loader_A,
    path_B: str, loader_B,
    cache_name_A: str = "ref_A_normalized.json",
    cache_name_B: str = "ref_B_normalized.json",
    force: bool = False,
) -> tuple[list[RequirementNormalized], list[RequirementNormalized]]:
    ref_A = _load_or_parse(path_A, loader_A, DATA_DIR / cache_name_A, force)
    ref_B = _load_or_parse(path_B, loader_B, DATA_DIR / cache_name_B, force)
    _validate_ingestion(Path(path_A).stem, ref_A)
    _validate_ingestion(Path(path_B).stem, ref_B)
    print(f"[Étape 1] Ref A : {len(ref_A)} exigences  |  Ref B : {len(ref_B)} exigences")
    return ref_A, ref_B


def run_ingestion_all(
    paths: list[str],
    force: bool = False,
) -> dict[str, list[RequirementNormalized]]:
    """Charge N référentiels en un seul appel.

    Args:
        paths: liste de chemins YAML à charger.
        force: si True, re-parse même si le cache existe.

    Returns:
        dict {fw_key: [RequirementNormalized, ...]} — une entrée par référentiel.
    """
    print(f"[Étape 1] Chargement de {len(paths)} référentiels ...")
    frameworks: dict[str, list[RequirementNormalized]] = {}
    for path in paths:
        fw_key = _name_from_path(Path(path))
        cache  = DATA_DIR / f"{re.sub(r'[^a-z0-9]', '_', fw_key.lower())}_normalized.json"
        reqs   = _load_or_parse(path, load_generic_yaml, cache, force)
        _validate_ingestion(Path(path).stem, reqs)
        frameworks[fw_key] = reqs
        print(f"  {fw_key:30s} : {len(reqs)} exigences")
    total = sum(len(v) for v in frameworks.values())
    print(f"[Étape 1] {len(frameworks)} référentiels — {total} exigences au total")
    return frameworks


# ══════════════════════════════════════════════════════════════════════════════
# Étape 2 — Similarité sémantique
# ══════════════════════════════════════════════════════════════════════════════

CACHE_PAIRS  = DATA_DIR / "candidate_pairs.json"
CACHE_MATRIX = DATA_DIR / "similarity_matrix.npy"


def _cache_slug(fw_a: str, fw_b: str) -> str:
    """Génère un slug stable pour nommer les caches par paire de frameworks."""
    import re as _re
    def slug(s: str) -> str:
        return _re.sub(r"[^a-z0-9]", "_", s.lower())[:30]
    return f"{slug(fw_a)}__{slug(fw_b)}"


def _get_embedding_model() -> SentenceTransformer:
    if not hasattr(_get_embedding_model, "_model"):
        try:
            import torch
            if torch.cuda.is_available():
                device = "cuda"
            elif hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
                device = "mps"
            else:
                device = "cpu"
        except ImportError:
            device = "cpu"
        print(f"  Chargement du modèle d'embedding ({EMBEDDING_MODEL}) sur {device.upper()}...")
        _get_embedding_model._model = SentenceTransformer(EMBEDDING_MODEL, device=device)
    return _get_embedding_model._model


def _req_text(r: RequirementNormalized) -> str:
    """Texte pour l'embedding : contexte parent (tags) | titre: description.
    
    Inspiré de CISO Assistant sbert_mapper: agrège le contexte hiérarchique
    pour réduire les faux négatifs sur vocabulaire divergent.
    """
    parts = []
    if r.tags:
        parts.append(r.tags[-1])  # contexte du parent immédiat
    first_sentence = r.description.split(".")[0].strip()[:200] if r.description else ""
    content = f"{r.title}: {first_sentence}" if first_sentence else r.title
    parts.append(content)
    return " | ".join(parts)


def _select_pairs(
    ref_A: list[RequirementNormalized],
    ref_B: list[RequirementNormalized],
    matrix: np.ndarray,
    semantic_threshold: float = SEMANTIC_THRESHOLD,
    top_k: int = TOP_K,
) -> list[CandidatePair]:
    pairs = []
    seen  = set()
    for i, req_a in enumerate(ref_A):
        scores      = matrix[i]
        top_indices = np.argsort(scores)[::-1][:top_k]
        for j in top_indices:
            score = float(scores[j])
            if score < semantic_threshold:
                break
            key = (req_a.id, ref_B[j].id)
            if key not in seen:
                seen.add(key)
                pairs.append(CandidatePair(
                    id_A=req_a.id,
                    title_A=req_a.title,
                    id_B=ref_B[j].id,
                    title_B=ref_B[j].title,
                    semantic_score=score,
                ))
    pairs.sort(key=lambda p: p.semantic_score, reverse=True)
    return pairs


def _save_pairs(pairs: list[CandidatePair], path: Path | None = None) -> None:
    out = path or CACHE_PAIRS
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump([p.to_dict() for p in pairs], f, ensure_ascii=False, indent=2)


def _load_pairs(path: Path | None = None) -> list[CandidatePair]:
    src = path or CACHE_PAIRS
    with open(src, "r", encoding="utf-8") as f:
        return [CandidatePair(**d) for d in json.load(f)]



def run_similarity(
    ref_A: list[RequirementNormalized],
    ref_B: list[RequirementNormalized],
    force: bool = False,
    semantic_threshold: float = SEMANTIC_THRESHOLD,
    top_k: int = TOP_K,
    fw_a_name: str = "",
    fw_b_name: str = "",
) -> tuple[list[CandidatePair], np.ndarray]:
    slug = _cache_slug(fw_a_name or "refA", fw_b_name or "refB")
    cache_pairs  = DATA_DIR / f"candidate_pairs_{slug}.json"
    cache_matrix = DATA_DIR / f"similarity_matrix_{slug}.npy"

    if cache_pairs.exists() and cache_matrix.exists() and not force:
        print("[Étape 2] Cache trouvé — rechargement des paires candidates.")
        pairs  = _load_pairs(cache_pairs)
        matrix = np.load(str(cache_matrix))
        print(f"  {len(pairs)} paires candidates chargées depuis le cache.")
        return pairs, matrix

    print(f"[Étape 2] Encodage de {len(ref_A)} + {len(ref_B)} textes avec '{EMBEDDING_MODEL}' ...")
    model = _get_embedding_model()

    emb_A = model.encode([_req_text(r) for r in ref_A], show_progress_bar=True, convert_to_numpy=True)
    emb_B = model.encode([_req_text(r) for r in ref_B], show_progress_bar=True, convert_to_numpy=True)

    emb_A = emb_A / np.linalg.norm(emb_A, axis=1, keepdims=True)
    emb_B = emb_B / np.linalg.norm(emb_B, axis=1, keepdims=True)

    matrix = emb_A @ emb_B.T
    pairs  = _select_pairs(ref_A, ref_B, matrix, semantic_threshold, top_k)

    np.save(str(cache_matrix), matrix)
    _save_pairs(pairs, cache_pairs)

    total = len(ref_A) * len(ref_B)
    print(f"  {len(pairs)} paires candidates retenues (seuil={semantic_threshold}, top_k={top_k}).")
    print(f"  Reduction : {total} combinaisons -> {len(pairs)} paires ({100 * len(pairs) / total:.1f}%)")
    return pairs, matrix


def run_similarity_all(
    frameworks: dict[str, list[RequirementNormalized]],
    force: bool = False,
    semantic_threshold: float = SEMANTIC_THRESHOLD,
    top_k: int = TOP_K,
) -> tuple[dict[tuple[str, str], list[CandidatePair]], dict[tuple[str, str], np.ndarray]]:
    """Encode toutes les exigences en UN seul passage, puis calcule les similarités
    pour chaque paire (fw_i, fw_j) avec i < j.

    Returns:
        candidates_all : dict[(fw_i, fw_j), list[CandidatePair]]
        matrices_all   : dict[(fw_i, fw_j), np.ndarray]
    """
    from itertools import combinations

    fw_keys = list(frameworks.keys())
    print(f"[Étape 2] Encodage global — {sum(len(v) for v in frameworks.values())} exigences "
          f"sur {len(fw_keys)} référentiels ...")

    model = _get_embedding_model()

    # Un seul encode() pour tout — embeddings indexés par fw_key
    emb_by_fw: dict[str, np.ndarray] = {}
    for fw_key, reqs in frameworks.items():
        raw = model.encode([_req_text(r) for r in reqs],
                           batch_size=256, show_progress_bar=False, convert_to_numpy=True)
        norms = np.linalg.norm(raw, axis=1, keepdims=True)
        emb_by_fw[fw_key] = raw / np.where(norms == 0, 1, norms)

    candidates_all: dict[tuple[str, str], list[CandidatePair]] = {}
    matrices_all:   dict[tuple[str, str], np.ndarray]          = {}

    pairs_iter = list(combinations(fw_keys, 2))
    print(f"  {len(pairs_iter)} paires à traiter ...")

    for fw_i, fw_j in pairs_iter:
        slug         = _cache_slug(fw_i, fw_j)
        cache_pairs  = DATA_DIR / f"candidate_pairs_{slug}.json"
        cache_matrix = DATA_DIR / f"similarity_matrix_{slug}.npy"

        if cache_pairs.exists() and cache_matrix.exists() and not force:
            candidates_all[(fw_i, fw_j)] = _load_pairs(cache_pairs)
            matrices_all[(fw_i, fw_j)]   = np.load(str(cache_matrix))
            print(f"  [{fw_i} <-> {fw_j}] cache -- {len(candidates_all[(fw_i, fw_j)])} paires")
            continue

        matrix = emb_by_fw[fw_i] @ emb_by_fw[fw_j].T
        pairs  = _select_pairs(frameworks[fw_i], frameworks[fw_j], matrix,
                                semantic_threshold, top_k)
        np.save(str(cache_matrix), matrix)
        _save_pairs(pairs, cache_pairs)

        total_comb = len(frameworks[fw_i]) * len(frameworks[fw_j])
        print(f"  [{fw_i} <-> {fw_j}] {len(pairs)}/{total_comb} paires "
              f"({100 * len(pairs) / max(total_comb, 1):.1f}%)")

        candidates_all[(fw_i, fw_j)] = pairs
        matrices_all[(fw_i, fw_j)]   = matrix

    total_pairs = sum(len(v) for v in candidates_all.values())
    print(f"[Étape 2] Total : {total_pairs} paires candidates sur {len(pairs_iter)} combinaisons")
    return candidates_all, matrices_all


# ══════════════════════════════════════════════════════════════════════════════
# Étape 3 — Heatmap
# ══════════════════════════════════════════════════════════════════════════════

def run_heatmap(
    ref_A: list[RequirementNormalized],
    ref_B: list[RequirementNormalized],
    matrix: np.ndarray,
    output_path: Path | None = None,
    show: bool = True,
) -> Path:
    out = output_path or OUTPUT_DIR / "correlation_matrix.png"
    out.parent.mkdir(parents=True, exist_ok=True)

    labels_A = [r.id for r in ref_A]
    labels_B = [r.id for r in ref_B]

    fig_w = max(20, len(labels_B) * 0.25)
    fig_h = max(14, len(labels_A) * 0.15)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    sns.heatmap(
        matrix,
        xticklabels=labels_B,
        yticklabels=labels_A,
        cmap="YlOrRd",
        vmin=0.0, vmax=1.0,
        ax=ax,
        linewidths=0,
        cbar_kws={"label": "Similarité cosinus"},
    )
    ax.set_title("Matrice de similarité sémantique — titre + description", fontsize=14, pad=12)
    ax.set_xlabel("Ref B", fontsize=11)
    ax.set_ylabel("Ref A", fontsize=11)
    ax.tick_params(axis="x", labelsize=6, rotation=90)
    ax.tick_params(axis="y", labelsize=6)

    plt.tight_layout()
    plt.savefig(str(out), dpi=150, bbox_inches="tight")
    if not show:
        plt.close(fig)

    print(f"[Etape 3] Heatmap sauvegardee -> {out}")
    return out


def run_heatmaps_all(
    frameworks: dict[str, list[RequirementNormalized]],
    matrices_all: dict[tuple[str, str], np.ndarray],
    show: bool = False,
) -> list[Path]:
    """Génère une heatmap pour chaque paire (fw_i, fw_j)."""
    paths = []
    for (fw_i, fw_j), matrix in matrices_all.items():
        slug = _cache_slug(fw_i, fw_j)
        out  = OUTPUT_DIR / f"heatmap_{slug}.png"
        # Titres d'axes avec le nom du framework
        ref_i = frameworks[fw_i]
        ref_j = frameworks[fw_j]
        path  = run_heatmap(ref_i, ref_j, matrix, output_path=out, show=show)
        paths.append(path)
    return paths


# ══════════════════════════════════════════════════════════════════════════════
# Étape 4 — LLM Scorer
# ══════════════════════════════════════════════════════════════════════════════

CACHE_RELATIONS = OUTPUT_DIR / "mapping_relations.json"


class LLMScoringOutput(BaseModel):
    coverage_A_to_B: float = Field(ge=0.0, le=1.0)
    coverage_B_to_A: float = Field(ge=0.0, le=1.0)
    confidence:      float = Field(ge=0.0, le=1.0)
    relation_type:   Literal["equivalence", "A_couvre_B", "B_couvre_A", "partielle", "aucun_lien"]
    justification:   str   = Field(default="", description="Explication courte du choix (1-2 phrases)")


class LLMBatchOutput(BaseModel):
    results: list[LLMScoringOutput]


PROMPT_INSTRUCTIONS = f"""Tu es un expert en mapping de référentiels de cybersécurité et conformité (ISO 27001, NIST CSF, CIS Controls, NIS2, DORA, SOC2, etc.).

## Définitions

**coverage_A_to_B** (float 0.0–1.0) : proportion des objectifs de sécurité de B qui sont couverts ou satisfaits par A.
- 1.0 = A adresse entièrement tous les objectifs de B
- 0.8 = A couvre la majorité des objectifs de B, quelques lacunes mineures
- 0.5 = A couvre environ la moitié des objectifs de B
- 0.2 = A effleure le sujet de B mais ne le couvre pas vraiment
- 0.0 = aucun rapport

**coverage_B_to_A** : idem mais dans le sens B→A.

**confidence** : ta certitude dans cette évaluation (0=incertain, 1=très certain).

**justification** : 1 à 2 phrases expliquant pourquoi tu as attribué ces scores et ce type de relation. Sois factuel et précis (ex : "A impose l'inventaire des actifs, B demande uniquement leur classification — A couvre donc B mais pas l'inverse.").

## Règles de classification relation_type

| Condition | relation_type |
|-----------|--------------|
| coverage_A_to_B >= {THRESHOLD_EQUIVALENCE} ET coverage_B_to_A >= {THRESHOLD_EQUIVALENCE} | "equivalence" |
| coverage_A_to_B >= {THRESHOLD_COVERAGE} ET coverage_B_to_A < {THRESHOLD_COVERAGE} | "A_couvre_B" |
| coverage_B_to_A >= {THRESHOLD_COVERAGE} ET coverage_A_to_B < {THRESHOLD_COVERAGE} | "B_couvre_A" |
| max(coverage_A_to_B, coverage_B_to_A) >= 0.4 | "partielle" |
| sinon | "aucun_lien" |

## Cas particuliers

- Si les deux exigences sont **quasi-identiques** (même titre, même sujet) → coverage_A_to_B=1.0, coverage_B_to_A=1.0, relation_type="equivalence", confidence=0.95, justification="Les deux exigences adressent le même objectif de sécurité."
- Tiens compte du **domaine** : une exigence de gestion des accès ne couvre pas une exigence de sauvegarde, même si les titres sont proches sémantiquement."""


def _make_pair_text(idx: int, req_a: RequirementNormalized, req_b: RequirementNormalized) -> str:
    tags_a = f" [{', '.join(req_a.tags[:3])}]" if req_a.tags else ""
    tags_b = f" [{', '.join(req_b.tags[:3])}]" if req_b.tags else ""
    desc_a = (req_a.description[:300] + "…") if len(req_a.description) > 300 else req_a.description
    desc_b = (req_b.description[:300] + "…") if len(req_b.description) > 300 else req_b.description
    return (
        f"--- Paire {idx} ---\n"
        f"A ({req_a.framework}{tags_a}) : {req_a.title}\n"
        f"{desc_a or '(description non disponible)'}\n\n"
        f"B ({req_b.framework}{tags_b}) : {req_b.title}\n"
        f"{desc_b or '(description non disponible)'}"
    )


def _build_prompt(pairs: list[tuple[RequirementNormalized, RequirementNormalized]]) -> str:
    """Construit un prompt unique (instructions + données) pour 1 ou N paires."""
    blocks = [_make_pair_text(i + 1, a, b) for i, (a, b) in enumerate(pairs)]
    pairs_text = "\n\n".join(blocks)

    if len(pairs) == 1:
        return (
            f"{PROMPT_INSTRUCTIONS}\n\n"
            f"## Paire à évaluer\n\n"
            f"{pairs_text}\n\n"
            f"## Format de réponse\n\n"
            f"Réponds UNIQUEMENT en JSON valide, sans commentaire, sans markdown.\n"
            f"Clés attendues : coverage_A_to_B, coverage_B_to_A, confidence, relation_type, justification."
        )
    else:
        return (
            f"{PROMPT_INSTRUCTIONS}\n\n"
            f"## Paires à évaluer ({len(pairs)})\n\n"
            f"{pairs_text}\n\n"
            f"## Format de réponse\n\n"
            f"Réponds UNIQUEMENT en JSON valide, sans commentaire, sans markdown.\n"
            f"Format attendu : {{\"results\": [<scoring_paire_1>, ..., <scoring_paire_{len(pairs)}>]}}\n"
            f"Chaque scoring_paire contient : coverage_A_to_B, coverage_B_to_A, confidence, relation_type, justification.\n"
            f"Le tableau \"results\" doit contenir exactement {len(pairs)} éléments."
        )


async def _score_batch(
    client: AsyncOpenAI,
    semaphore: asyncio.Semaphore,
    pairs: list[tuple[RequirementNormalized, RequirementNormalized]],
    llm_model: str,
) -> list[LLMScoringOutput | None]:
    """Score un batch de paires en un seul appel API. Fallback 1-par-1 si parse échoue."""
    if not pairs:
        return []

    prompt = _build_prompt(pairs)

    async with semaphore:
        for attempt in range(LLM_MAX_RETRIES + 1):
            try:
                response = await client.chat.completions.create(
                    model=llm_model,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0.0,
                )
                raw = response.choices[0].message.content

                if len(pairs) == 1:
                    result = LLMScoringOutput.model_validate_json(raw)
                    return [result]

                batch_out = LLMBatchOutput.model_validate_json(raw)
                if len(batch_out.results) != len(pairs):
                    raise ValueError(
                        f"Batch: attendu {len(pairs)} résultats, reçu {len(batch_out.results)}"
                    )
                return batch_out.results

            except Exception as e:
                if attempt == LLM_MAX_RETRIES:
                    if len(pairs) > 1:
                        # Fallback : scorer 1 par 1
                        print(f"  [LLM BATCH FALLBACK] {len(pairs)} paires → 1-par-1 ({e})")
                        results = []
                        for a, b in pairs:
                            r = await _score_batch(client, semaphore, [(a, b)], llm_model)
                            results.append(r[0] if r else None)
                        return results
                    else:
                        print(f"  [LLM ERREUR] {pairs[0][0].id}↔{pairs[0][1].id} : {e}")
                        return [None]
    return [None] * len(pairs)


async def _run_scorer_async(
    candidates: list[CandidatePair],
    ref_A_map: dict[str, RequirementNormalized],
    ref_B_map: dict[str, RequirementNormalized],
    llm_model: str,
    batch_size: int,
) -> list[MappingRelation]:
    client    = AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))
    semaphore = asyncio.Semaphore(LLM_CONCURRENCY)

    # Découper les candidats en batches
    pair_objects = [(ref_A_map[p.id_A], ref_B_map[p.id_B]) for p in candidates]
    batches      = [pair_objects[i:i + batch_size] for i in range(0, len(pair_objects), batch_size)]
    cand_batches = [candidates[i:i + batch_size]   for i in range(0, len(candidates),   batch_size)]

    n_calls = len(batches)
    print(
        f"[Étape 4] Scoring LLM sur {len(candidates)} paires "
        f"({n_calls} appels API, batch_size={batch_size}) ..."
    )

    batch_tasks = [
        _score_batch(client, semaphore, batch, llm_model)
        for batch in batches
    ]
    batch_results = await asyncio.gather(*batch_tasks)

    # Aplatir résultats
    relations: list[MappingRelation] = []
    for cand_batch, results in zip(cand_batches, batch_results):
        for pair, result in zip(cand_batch, results):
            if result is None:
                continue
            cov_a = result.coverage_A_to_B
            cov_b = result.coverage_B_to_A
            relations.append(MappingRelation(
                id_A=pair.id_A, title_A=pair.title_A,
                id_B=pair.id_B, title_B=pair.title_B,
                semantic_score=pair.semantic_score,
                coverage_A_to_B=cov_a,
                coverage_B_to_A=cov_b,
                confidence=result.confidence,
                relation_type=_infer_relation_type(cov_a, cov_b),
                justification=result.justification,
            ))

    # Double vérification sur les paires à fort score
    high_score = [
        r for r in relations
        if r.coverage_A_to_B >= LLM_CONFIRM_THRESHOLD or r.coverage_B_to_A >= LLM_CONFIRM_THRESHOLD
    ]
    if high_score:
        print(f"  Double vérification sur {len(high_score)} paires à fort score ...")
        confirm_pairs   = [(ref_A_map[r.id_A], ref_B_map[r.id_B]) for r in high_score]
        confirm_batches = [confirm_pairs[i:i + batch_size] for i in range(0, len(confirm_pairs), batch_size)]
        confirm_tasks   = [_score_batch(client, semaphore, b, llm_model) for b in confirm_batches]
        confirm_results_nested = await asyncio.gather(*confirm_tasks)
        confirm_results = [r for batch in confirm_results_nested for r in batch]

        confirmed_map = {(r.id_A, r.id_B): r for r in high_score}
        for rel, confirm in zip(high_score, confirm_results):
            if confirm is None:
                continue
            orig = confirmed_map[(rel.id_A, rel.id_B)]
            cov_a = round((orig.coverage_A_to_B + confirm.coverage_A_to_B) / 2, 4)
            cov_b = round((orig.coverage_B_to_A + confirm.coverage_B_to_A) / 2, 4)
            orig.coverage_A_to_B = cov_a
            orig.coverage_B_to_A = cov_b
            orig.confidence      = round((orig.confidence + confirm.confidence) / 2, 4)
            orig.relation_type   = _infer_relation_type(cov_a, cov_b)
            # Conserver la justification la plus détaillée (passe 2 si disponible)
            if confirm.justification:
                orig.justification = confirm.justification

    return relations


def _export_results(
    relations: list[MappingRelation],
    removed: list[MappingRelation] | None = None,
    xlsx_path: Path | None = None,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with open(CACHE_RELATIONS, "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in relations], f, ensure_ascii=False, indent=2)
    print(f"  → {CACHE_RELATIONS}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mapping Results"

        headers = [
            "ID Ref A", "Titre Ref A", "ID Ref B", "Titre Ref B",
            "Score sémantique", "Coverage A→B", "Coverage B→A",
            "Confiance", "Type de relation", "Justification",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = border

        RELATION_COLORS = {
            "equivalence": "C6EFCE",
            "A_couvre_B":  "FFEB9C",
            "B_couvre_A":  "FFEB9C",
            "partielle":   "FFDDC1",
            "aucun_lien":  "FFC7CE",
        }

        for row_idx, r in enumerate(relations, 2):
            values = [
                r.id_A, r.title_A, r.id_B, r.title_B,
                r.semantic_score, r.coverage_A_to_B, r.coverage_B_to_A,
                r.confidence, r.relation_type, r.justification,
            ]
            row_fill = PatternFill("solid", fgColor=RELATION_COLORS.get(r.relation_type, "FFFFFF"))
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_idx in (5, 6, 7, 8):
                    cell.number_format = "0.00"
                if col_idx == 9:
                    cell.fill = row_fill

        col_widths = [12, 50, 16, 50, 12, 12, 12, 10, 18, 60]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Feuille "Supprimées" (aucun_lien filtrés)
        if removed:
            ws2 = wb.create_sheet("Supprimées")
            for col, h in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.fill = PatternFill("solid", fgColor="808080")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.border = border
            for row_idx, r in enumerate(removed, 2):
                values = [
                    r.id_A, r.title_A, r.id_B, r.title_B,
                    r.semantic_score, r.coverage_A_to_B, r.coverage_B_to_A,
                    r.confidence, r.relation_type, r.justification,
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    if col_idx in (5, 6, 7, 8):
                        cell.number_format = "0.00"

        out = xlsx_path or OUTPUT_DIR / "mapping_results.xlsx"
        wb.save(str(out))
        print(f"  → {out}")
        html_out = out.with_suffix(".html")
        _export_html_review(relations, html_out)

    except Exception as e:
        print(f"  [WARN] Export Excel échoué : {e}")


def run_scorer(
    candidates: list[CandidatePair],
    ref_A: list[RequirementNormalized],
    ref_B: list[RequirementNormalized],
    force: bool = False,
    llm_model: str = LLM_MODEL,
    batch_size: int = LLM_BATCH_SIZE,
    cache_path: Path | None = None,
) -> list[MappingRelation]:
    _cache = cache_path or CACHE_RELATIONS
    if _cache.exists() and not force:
        print("[Étape 4] Cache trouvé — rechargement des relations.")
        with open(_cache, "r", encoding="utf-8") as f:
            data = json.load(f)
        relations = [
            MappingRelation(**{k: v for k, v in d.items() if k in MappingRelation.__dataclass_fields__})
            for d in data
        ]
        print(f"  {len(relations)} relations chargées depuis le cache.")
        return relations

    ref_A_map = {r.id: r for r in ref_A}
    ref_B_map = {r.id: r for r in ref_B}

    relations = asyncio.run(
        _run_scorer_async(candidates, ref_A_map, ref_B_map, llm_model, batch_size)
    )

    print(f"\n[Diagnostic étape 4]")
    print(f"  Paires scorées : {len(relations)}")
    for rt in ["equivalence", "A_couvre_B", "B_couvre_A", "partielle", "aucun_lien"]:
        n = sum(1 for r in relations if r.relation_type == rt)
        print(f"  {rt:20s} : {n}")

    # Sauvegarder dans le bon cache
    _cache.parent.mkdir(parents=True, exist_ok=True)
    with open(_cache, "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in relations], f, ensure_ascii=False, indent=2)

    return relations


# ══════════════════════════════════════════════════════════════════════════════
# Étape 5 — Nettoyage des liaisons inutiles
# ══════════════════════════════════════════════════════════════════════════════

def run_cleanup(
    relations: list[MappingRelation],
    min_confidence: float = LLM_MIN_CONFIDENCE,
) -> tuple[list[MappingRelation], list[MappingRelation]]:
    """
    Filtre les relations sans valeur :
    - relation_type == "aucun_lien"
    - confidence < min_confidence

    Retourne (relations_utiles, relations_supprimées).
    """
    clean, removed = [], []
    for r in relations:
        if r.relation_type == "aucun_lien" or r.confidence < min_confidence:
            removed.append(r)
        else:
            clean.append(r)

    print(f"[Étape 5] Nettoyage : {len(clean)} relations conservées, {len(removed)} supprimées "
          f"(aucun_lien ou confiance < {min_confidence:.2f})")
    return clean, removed


def _favor_equivalence_filter(
    relations: list[MappingRelation],
) -> list[MappingRelation]:
    """Pour chaque id_A, si des équivalences existent, ne garder qu'elles.

    Inspiré de CISO Assistant favor_equals_filter : réduit le bruit en
    priorisant les correspondances fortes quand elles existent.
    """
    from collections import defaultdict
    by_source: dict[str, list[MappingRelation]] = defaultdict(list)
    for r in relations:
        by_source[r.id_A].append(r)
    result = []
    for rels in by_source.values():
        equiv = [r for r in rels if r.relation_type == "equivalence"]
        result.extend(equiv if equiv else rels)
    return result


def _export_html_review(
    relations: list[MappingRelation],
    output_path: Path,
) -> None:
    """Génère un fichier HTML interactif pour la revue humaine des mappings.

    Inspiré de CISO Assistant prepare_review.py : checkboxes de validation,
    barre de progression, persistance localStorage, couleurs par type de relation.
    """
    RELATION_CSS = {
        "equivalence": "#27ae60",
        "A_couvre_B":  "#3498db",
        "B_couvre_A":  "#e67e22",
        "partielle":   "#f39c12",
        "aucun_lien":  "#95a5a6",
    }

    import html as _html

    rows_html = []
    for idx, r in enumerate(relations):
        color = RELATION_CSS.get(r.relation_type, "#999")
        e_id_a   = _html.escape(r.id_A)
        e_title_a = _html.escape(r.title_A)
        e_rel    = _html.escape(r.relation_type)
        e_id_b   = _html.escape(r.id_B)
        e_title_b = _html.escape(r.title_B)
        e_just   = _html.escape(r.justification or "")
        rows_html.append(f"""
        <tr>
          <td class="ref-id">{e_id_a}</td>
          <td class="desc">{e_title_a}</td>
          <td><span class="badge" style="background:{color}">{e_rel}</span><br>
              <small style="color:#666">{r.semantic_score:.2f} → {r.coverage_A_to_B:.2f}/{r.coverage_B_to_A:.2f}</small></td>
          <td class="ref-id">{e_id_b}</td>
          <td class="desc">{e_title_b}</td>
          <td class="desc" style="font-style:italic;color:#555">{e_just}</td>
          <td style="text-align:center"><input type="checkbox" id="cb_{idx}" onchange="updateStats()"></td>
        </tr>""")

    rows_joined = "\n".join(rows_html)
    total = len(relations)

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>RiskHunter — Mapping Review</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f4f6f8; padding: 20px; color: #333; }}
  .container {{ max-width: 1600px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,.1); padding: 30px; }}
  h1 {{ color: #1a2e4a; margin-bottom: 8px; }}
  .subtitle {{ color: #666; font-size: .9em; margin-bottom: 20px; }}
  .progress-wrap {{ background: #e0e0e0; border-radius: 4px; height: 28px; overflow: hidden; margin-bottom: 20px; }}
  .progress-bar {{ height: 100%; background: linear-gradient(90deg,#3498db,#27ae60); display:flex; align-items:center; justify-content:center; color:#fff; font-weight:700; transition: width .3s; min-width: 40px; }}
  .stats {{ display:flex; gap:16px; margin-bottom:20px; }}
  .stat {{ flex:1; background:#f0f4f8; border-radius:6px; padding:12px; text-align:center; }}
  .stat-val {{ font-size:2em; font-weight:700; color:#3498db; }}
  .stat-lbl {{ font-size:.8em; color:#888; }}
  table {{ width:100%; border-collapse:collapse; font-size:.85em; }}
  thead {{ background:#1a2e4a; color:#fff; position:sticky; top:0; }}
  th {{ padding:10px 8px; text-align:left; }}
  td {{ padding:8px; border-bottom:1px solid #eee; vertical-align:top; }}
  tr:hover {{ background:#fafafa; }}
  tr.validated {{ background:#eafaf1; }}
  .ref-id {{ font-family:monospace; font-weight:600; white-space:nowrap; }}
  .desc {{ max-width:280px; line-height:1.4; }}
  .badge {{ display:inline-block; padding:3px 8px; border-radius:4px; color:#fff; font-size:.8em; font-weight:600; }}
  .legend {{ display:flex; gap:12px; flex-wrap:wrap; margin-bottom:16px; }}
  .legend-item {{ display:flex; align-items:center; gap:6px; font-size:.8em; }}
</style>
</head>
<body>
<div class="container">
  <h1>🔍 RiskHunter — Mapping Review</h1>
  <p class="subtitle">{total} relations à valider</p>

  <div class="legend">
    <span class="legend-item"><span class="badge" style="background:#27ae60">equivalence</span> Équivalence complète</span>
    <span class="legend-item"><span class="badge" style="background:#3498db">A_couvre_B</span> A couvre B</span>
    <span class="legend-item"><span class="badge" style="background:#e67e22">B_couvre_A</span> B couvre A</span>
    <span class="legend-item"><span class="badge" style="background:#f39c12">partielle</span> Couverture partielle</span>
    <span class="legend-item"><span class="badge" style="background:#95a5a6">aucun_lien</span> Aucun lien</span>
  </div>

  <div class="progress-wrap">
    <div class="progress-bar" id="progressBar" style="width:0%">0%</div>
  </div>

  <div class="stats">
    <div class="stat"><div class="stat-val">{total}</div><div class="stat-lbl">Total</div></div>
    <div class="stat"><div class="stat-val" id="reviewedCount">0</div><div class="stat-lbl">Validées</div></div>
    <div class="stat"><div class="stat-val" id="remainingCount">{total}</div><div class="stat-lbl">Restantes</div></div>
  </div>

  <table>
    <thead>
      <tr>
        <th>ID A</th><th>Titre A</th><th>Relation</th>
        <th>ID B</th><th>Titre B</th><th>Justification</th><th>✓</th>
      </tr>
    </thead>
    <tbody>
{rows_joined}
    </tbody>
  </table>
</div>
<script>
window._rhReportKey = {repr(str(output_path.name))};
function updateStats() {{
  const cbs = document.querySelectorAll('input[type=checkbox]');
  const total = cbs.length;
  const done = Array.from(cbs).filter(c => c.checked).length;
  document.getElementById('reviewedCount').textContent = done;
  document.getElementById('remainingCount').textContent = total - done;
  const pct = total ? Math.round(done / total * 100) : 0;
  const bar = document.getElementById('progressBar');
  bar.style.width = pct + '%';
  bar.textContent = pct + '%';
  cbs.forEach(cb => cb.closest('tr').classList.toggle('validated', cb.checked));
  localStorage.setItem('rh_review_state_' + window._rhReportKey, JSON.stringify(Array.from(cbs).map(c => c.checked)));
}}
(function() {{
  const saved = localStorage.getItem('rh_review_state_' + window._rhReportKey);
  if (!saved) return;
  const state = JSON.parse(saved);
  const cbs = document.querySelectorAll('input[type=checkbox]');
  state.forEach((v, i) => {{ if (cbs[i]) cbs[i].checked = v; }});
  updateStats();
}})();
</script>
</body>
</html>"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    print(f"  → {output_path}")


def run_scorer_all(
    candidates_all: dict[tuple[str, str], list[CandidatePair]],
    frameworks: dict[str, list[RequirementNormalized]],
    force: bool = False,
    llm_model: str = LLM_MODEL,
    batch_size: int = LLM_BATCH_SIZE,
) -> dict[tuple[str, str], list[MappingRelation]]:
    """Scorage LLM pour toutes les paires (fw_i, fw_j).

    Injecte framework_A et framework_B dans chaque MappingRelation produite.
    Cache par paire dans OUTPUT_DIR/mapping_relations_{slug}.json.
    """
    relations_all: dict[tuple[str, str], list[MappingRelation]] = {}

    for (fw_i, fw_j), candidates in candidates_all.items():
        slug      = _cache_slug(fw_i, fw_j)
        cache_rel = OUTPUT_DIR / f"mapping_relations_{slug}.json"
        print(f"\n[Etape 4] {fw_i} <-> {fw_j} -- {len(candidates)} paires candidates ...")

        rels = run_scorer(
            candidates,
            frameworks[fw_i],
            frameworks[fw_j],
            force=force,
            llm_model=llm_model,
            batch_size=batch_size,
            cache_path=cache_rel,
        )
        # Injecter framework_A / framework_B
        for r in rels:
            r.framework_A = fw_i
            r.framework_B = fw_j

        relations_all[(fw_i, fw_j)] = rels

    total = sum(len(v) for v in relations_all.values())
    print(f"\n[Étape 4] Total : {total} relations sur {len(relations_all)} paires")
    return relations_all


def run_cleanup_all(
    relations_all: dict[tuple[str, str], list[MappingRelation]],
    min_confidence: float = LLM_MIN_CONFIDENCE,
) -> tuple[dict[tuple[str, str], list[MappingRelation]], list[MappingRelation]]:
    """Nettoyage pour toutes les paires. Retourne (clean_all, removed_flat)."""
    clean_all: dict[tuple[str, str], list[MappingRelation]] = {}
    removed_flat: list[MappingRelation] = []
    for pair, rels in relations_all.items():
        clean, removed = run_cleanup(rels, min_confidence=min_confidence)
        clean_all[pair] = clean
        removed_flat.extend(removed)
    return clean_all, removed_flat


def _export_global(
    relations_all: dict[tuple[str, str], list[MappingRelation]],
    removed_flat: list[MappingRelation] | None = None,
) -> None:
    """Exporte l'ensemble des relations (toutes paires confondues).

    Produit :
    - data/outputs/mapping_all.json        (toutes relations, auto-suffisant avec framework_A/B)
    - data/outputs/mapping_summary.xlsx    (1 feuille par paire + feuille "Global")
    - data/outputs/mapping_all.html        (revue interactive)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    all_relations = [r for rels in relations_all.values() for r in rels]

    # ── JSON global ──────────────────────────────────────────────────────────
    json_path = OUTPUT_DIR / "mapping_all.json"
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in all_relations], f, ensure_ascii=False, indent=2)
    print(f"[Export] JSON global -> {json_path} ({len(all_relations)} relations)")

    # ── HTML revue ───────────────────────────────────────────────────────────
    html_path = OUTPUT_DIR / "mapping_all.html"
    _export_html_review(all_relations, html_path)

    # ── Excel multi-feuilles ─────────────────────────────────────────────────
    try:
        wb = Workbook()
        wb.remove(wb.active)  # supprimer la feuille vide par défaut

        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        RELATION_COLORS = {
            "equivalence": "C6EFCE",
            "A_couvre_B":  "FFEB9C",
            "B_couvre_A":  "FFEB9C",
            "partielle":   "FFDDC1",
            "aucun_lien":  "FFC7CE",
        }
        HEADERS = [
            "Framework A", "ID Ref A", "Titre Ref A",
            "Framework B", "ID Ref B", "Titre Ref B",
            "Score sémantique", "Coverage A→B", "Coverage B→A",
            "Confiance", "Type de relation", "Justification",
        ]

        def _write_sheet(ws, relations_list: list[MappingRelation]) -> None:
            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            hdr_font = Font(bold=True, color="FFFFFF")
            for col, h in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = border
            for row_idx, r in enumerate(relations_list, 2):
                vals = [
                    r.framework_A, r.id_A, r.title_A,
                    r.framework_B, r.id_B, r.title_B,
                    r.semantic_score, r.coverage_A_to_B, r.coverage_B_to_A,
                    r.confidence, r.relation_type, r.justification,
                ]
                row_fill = PatternFill("solid", fgColor=RELATION_COLORS.get(r.relation_type, "FFFFFF"))
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col_idx in (7, 8, 9, 10):
                        cell.number_format = "0.00"
                    if col_idx == 11:
                        cell.fill = row_fill
            col_widths = [18, 12, 45, 18, 12, 45, 12, 12, 12, 10, 18, 60]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes    = "A2"
            ws.auto_filter.ref = ws.dimensions

        # Feuille par paire
        for (fw_i, fw_j), rels in relations_all.items():
            sheet_name = f"{fw_i[:12]} {fw_j[:12]}"[:31]  # Excel: max 31 chars
            _write_sheet(wb.create_sheet(sheet_name), rels)

        # Feuille globale
        _write_sheet(wb.create_sheet("Global"), all_relations)

        # Feuille Supprimées
        if removed_flat:
            _write_sheet(wb.create_sheet("Supprimées"), removed_flat)

        xlsx_path = OUTPUT_DIR / "mapping_summary.xlsx"
        wb.save(str(xlsx_path))
        print(f"[Export] Excel global -> {xlsx_path}")
    except Exception as e:
        print(f"  [WARN] Export Excel global échoué : {e}")


# ══════════════════════════════════════════════════════════════════════════════
# Pipeline principal
# ══════════════════════════════════════════════════════════════════════════════

def main():
    SURVEYS_DIR = Path(__file__).parent / "files" / "surveys"

    parser = argparse.ArgumentParser(description="Pipeline de mapping référentiels (standalone)")

    # Mode N-frameworks (nouveau)
    parser.add_argument("--frameworks", nargs="+", metavar="YAML",
                        help="Chemins vers N fichiers YAML (mode multi-frameworks)")
    parser.add_argument("--all", dest="all_frameworks", action="store_true",
                        help=f"Scanner tous les YAML dans {SURVEYS_DIR}")

    # Mode pairwise (rétrocompatibilité)
    parser.add_argument("--ref-a",         default=None, help="Chemin vers le fichier Ref A (YAML)")
    parser.add_argument("--adapter-a",     default="generic")
    parser.add_argument("--ref-b",         default=None, help="Chemin vers le fichier Ref B (YAML)")
    parser.add_argument("--adapter-b",     default="generic")

    # Communs
    parser.add_argument("--force-step1",   action="store_true")
    parser.add_argument("--force-step2",   action="store_true")
    parser.add_argument("--force-step4",   action="store_true")
    parser.add_argument("--skip-step4",    action="store_true", help="Arrêter après l'étape 3")
    parser.add_argument("--skip-step5",    action="store_true", help="Ne pas filtrer les aucun_lien")
    parser.add_argument("--min-confidence", type=float, default=LLM_MIN_CONFIDENCE,
                        help=f"Confiance minimale pour conserver une relation (défaut: {LLM_MIN_CONFIDENCE})")
    args = parser.parse_args()

    # ── Résoudre la liste de frameworks ──────────────────────────────────────
    if args.all_frameworks:
        yaml_paths = sorted(SURVEYS_DIR.glob("*.yaml")) + sorted(SURVEYS_DIR.glob("*.yml"))
        if not yaml_paths:
            sys.exit(f"Aucun fichier YAML trouvé dans {SURVEYS_DIR}")
        print(f"[Pipeline] Mode --all : {len(yaml_paths)} référentiels détectés")
    elif args.frameworks:
        yaml_paths = [Path(p) for p in args.frameworks]
    elif args.ref_a and args.ref_b:
        yaml_paths = None  # mode pairwise classique
    else:
        parser.error("Spécifiez --frameworks, --all, ou --ref-a/--ref-b")

    # ══════════════════════════════════════════════════════════════════════════
    # Mode N-frameworks
    # ══════════════════════════════════════════════════════════════════════════
    if yaml_paths is not None:
        # ── Étape 1 ──────────────────────────────────────────────────────────
        frameworks = run_ingestion_all(yaml_paths, force=args.force_step1)
        if len(frameworks) < 2:
            sys.exit("Il faut au moins 2 référentiels valides pour faire un mapping.")

        # ── Étape 2 ──────────────────────────────────────────────────────────
        candidates_all, matrices_all = run_similarity_all(
            frameworks, force=args.force_step2
        )

        # ── Étape 3 ──────────────────────────────────────────────────────────
        run_heatmaps_all(frameworks, matrices_all, show=False)

        if args.skip_step4:
            print("\n[Pipeline] Étape 4 ignorée (--skip-step4). Pipeline terminé.")
            return

        # ── Étape 4 ──────────────────────────────────────────────────────────
        relations_all = run_scorer_all(
            candidates_all, frameworks, force=args.force_step4
        )

        # ── Étape 5 ──────────────────────────────────────────────────────────
        removed_flat: list[MappingRelation] = []
        if not args.skip_step5:
            relations_all, removed_flat = run_cleanup_all(
                relations_all, min_confidence=args.min_confidence
            )

        # ── Export global ─────────────────────────────────────────────────────
        _export_global(relations_all, removed_flat=removed_flat or None)

        total = sum(len(v) for v in relations_all.values())
        print(f"\n[Pipeline] Terminé — {total} relations exportées "
              f"({len(removed_flat)} supprimées) sur {len(relations_all)} paires.")
        return

    # ══════════════════════════════════════════════════════════════════════════
    # Mode pairwise (rétrocompatibilité --ref-a / --ref-b)
    # ══════════════════════════════════════════════════════════════════════════
    loader_a = ADAPTER_LOADERS.get(args.adapter_a)
    loader_b = ADAPTER_LOADERS.get(args.adapter_b)
    if loader_a is None:
        sys.exit(f"Adaptateur inconnu : {args.adapter_a!r}. Disponibles : {list(ADAPTER_LOADERS)}")
    if loader_b is None:
        sys.exit(f"Adaptateur inconnu : {args.adapter_b!r}. Disponibles : {list(ADAPTER_LOADERS)}")

    fw_a = Path(args.ref_a).stem
    fw_b = Path(args.ref_b).stem
    slug = _cache_slug(fw_a, fw_b)

    # ── Étape 1 ──────────────────────────────────────────────────────────────
    ref_A, ref_B = run_ingestion(
        path_A=args.ref_a, loader_A=loader_a,
        path_B=args.ref_b, loader_B=loader_b,
        force=args.force_step1,
    )

    # ── Étape 2 ──────────────────────────────────────────────────────────────
    candidates, matrix = run_similarity(
        ref_A, ref_B,
        force=args.force_step2,
        fw_a_name=fw_a, fw_b_name=fw_b,
    )

    # ── Étape 3 ──────────────────────────────────────────────────────────────
    run_heatmap(ref_A, ref_B, matrix, show=False)

    if args.skip_step4:
        print("\n[Pipeline] Étape 4 ignorée (--skip-step4). Pipeline terminé.")
        return

    # ── Étape 4 ──────────────────────────────────────────────────────────────
    cache_rel = OUTPUT_DIR / f"mapping_relations_{slug}.json"
    relations = run_scorer(
        candidates, ref_A, ref_B,
        force=args.force_step4,
        cache_path=cache_rel,
    )

    # ── Étape 5 ──────────────────────────────────────────────────────────────
    if not args.skip_step5:
        relations, removed = run_cleanup(relations, min_confidence=args.min_confidence)
        xlsx_out = OUTPUT_DIR / f"mapping_results_{slug}.xlsx"
        _export_results(relations, removed=removed, xlsx_path=xlsx_out)
    else:
        xlsx_out = OUTPUT_DIR / f"mapping_results_{slug}.xlsx"
        _export_results(relations, xlsx_path=xlsx_out)
        removed = []

    print(f"\n[Pipeline] Terminé — {len(relations)} relations exportées ({len(removed)} supprimées).")


if __name__ == "__main__":
    main()
