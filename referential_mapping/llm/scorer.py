"""
llm/scorer.py — Étape 4 : vérification intelligente LLM (titre + description).

- Filtre les paires candidates au-dessus du seuil LLM_CONFIRM_THRESHOLD
- Appelle le LLM pour scorer chaque paire (coverage A→B, B→A, confidence, relation_type)
- Double vérification sur les paires avec score LLM très élevé
- Export final : mapping_results.xlsx + mapping_relations.json
"""
import asyncio
import json
import os
from pathlib import Path
from typing import Literal
import dotenv
dotenv.load_dotenv()

from openai import AsyncOpenAI
from pydantic import BaseModel, Field, ValidationError

from referential_mapping.models import RequirementNormalized, CandidatePair, MappingRelation
from referential_mapping.config import (
    OUTPUT_DIR, LLM_MODEL, LLM_MAX_RETRIES, LLM_CONFIRM_THRESHOLD, LLM_CONCURRENCY
)

CACHE_RELATIONS = OUTPUT_DIR / "mapping_relations.json"


# ── Schéma de sortie LLM ─────────────────────────────────────────────────────

class LLMScoringOutput(BaseModel):
    coverage_A_to_B: float = Field(ge=0.0, le=1.0)
    coverage_B_to_A: float = Field(ge=0.0, le=1.0)
    confidence:      float = Field(ge=0.0, le=1.0)
    relation_type:   Literal["equivalence", "A_couvre_B", "B_couvre_A", "partielle", "aucun_lien"]


# ── Prompt ───────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Tu es un expert en référentiels de cybersécurité et conformité.
Tu dois évaluer la corrélation entre deux exigences issues de référentiels différents.
Réponds UNIQUEMENT en JSON valide, sans commentaire, avec exactement ces clés :
- coverage_A_to_B (float 0-1) : dans quelle mesure A couvre/satisfait B
- coverage_B_to_A (float 0-1) : dans quelle mesure B couvre/satisfait A
- confidence (float 0-1) : ta confiance dans cette évaluation
- relation_type : exactement l'une de ces valeurs :
  "equivalence" | "A_couvre_B" | "B_couvre_A" | "partielle" | "aucun_lien"

Règles pour relation_type :
- equivalence  : coverage_A_to_B >= 0.8 ET coverage_B_to_A >= 0.8
- A_couvre_B   : coverage_A_to_B >= 0.8 ET coverage_B_to_A < 0.8
- B_couvre_A   : coverage_B_to_A >= 0.8 ET coverage_A_to_B < 0.8
- partielle    : coverage_A_to_B >= 0.4 OU coverage_B_to_A >= 0.4
- aucun_lien   : sinon"""


def _make_user_prompt(req_a: RequirementNormalized, req_b: RequirementNormalized) -> str:
    return f"""Exigence A ({req_a.framework}) :
Titre : {req_a.title}
Description : {req_a.description or "(non disponible)"}

Exigence B ({req_b.framework}) :
Titre : {req_b.title}
Description : {req_b.description or "(non disponible)"}

Évalue la corrélation entre ces deux exigences."""


# ── Appel LLM ────────────────────────────────────────────────────────────────

async def _score_pair(
    client: AsyncOpenAI,
    semaphore: asyncio.Semaphore,
    req_a: RequirementNormalized,
    req_b: RequirementNormalized,
) -> LLMScoringOutput | None:
    async with semaphore:
        for attempt in range(LLM_MAX_RETRIES + 1):
            try:
                response = await client.chat.completions.create(
                    model=LLM_MODEL,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user",   "content": _make_user_prompt(req_a, req_b)},
                    ],
                    temperature=0.0,
                )
                raw = response.choices[0].message.content
                return LLMScoringOutput.model_validate_json(raw)
            except (ValidationError, json.JSONDecodeError, Exception) as e:
                if attempt == LLM_MAX_RETRIES:
                    print(f"  [LLM ERREUR] {req_a.id}↔{req_b.id} : {e}")
                    return None
    return None


# ── Orchestration étape 4 ────────────────────────────────────────────────────

async def _run_async(
    candidates: list[CandidatePair],
    ref_A_map: dict[str, RequirementNormalized],
    ref_B_map: dict[str, RequirementNormalized],
) -> list[MappingRelation]:
    api_key = os.environ.get("OPENAI_API_KEY", "")
    client = AsyncOpenAI(api_key=api_key)
    semaphore = asyncio.Semaphore(LLM_CONCURRENCY)

    relations: list[MappingRelation] = []
    total = len(candidates)

    tasks = [
        _score_pair(client, semaphore, ref_A_map[p.id_A], ref_B_map[p.id_B])
        for p in candidates
    ]

    print(f"[Étape 4] Scoring LLM sur {total} paires candidates ...")
    results = await asyncio.gather(*tasks)

    for pair, result in zip(candidates, results):
        if result is None:
            continue
        relations.append(MappingRelation(
            id_A=pair.id_A,
            title_A=pair.title_A,
            id_B=pair.id_B,
            title_B=pair.title_B,
            semantic_score=pair.semantic_score,
            coverage_A_to_B=result.coverage_A_to_B,
            coverage_B_to_A=result.coverage_B_to_A,
            confidence=result.confidence,
            relation_type=result.relation_type,
        ))

    # Double vérification sur les paires à fort score LLM
    high_score = [r for r in relations if r.coverage_A_to_B >= LLM_CONFIRM_THRESHOLD
                  or r.coverage_B_to_A >= LLM_CONFIRM_THRESHOLD]
    if high_score:
        print(f"  Double vérification sur {len(high_score)} paires à fort score ...")
        confirm_tasks = [
            _score_pair(client, semaphore, ref_A_map[r.id_A], ref_B_map[r.id_B])
            for r in high_score
        ]
        confirm_results = await asyncio.gather(*confirm_tasks)

        # Moyenne des deux passes pour les paires confirmées
        confirmed_map = {(r.id_A, r.id_B): r for r in high_score}
        for rel, confirm in zip(high_score, confirm_results):
            if confirm is None:
                continue
            key = (rel.id_A, rel.id_B)
            orig = confirmed_map[key]
            orig.coverage_A_to_B = round((orig.coverage_A_to_B + confirm.coverage_A_to_B) / 2, 4)
            orig.coverage_B_to_A = round((orig.coverage_B_to_A + confirm.coverage_B_to_A) / 2, 4)
            orig.confidence       = round((orig.confidence + confirm.confidence) / 2, 4)

    return relations


def run(
    candidates: list[CandidatePair],
    ref_A: list[RequirementNormalized],
    ref_B: list[RequirementNormalized],
    force: bool = False,
) -> list[MappingRelation]:
    if CACHE_RELATIONS.exists() and not force:
        print("[Étape 4] Cache trouvé — rechargement des relations.")
        with open(CACHE_RELATIONS, "r", encoding="utf-8") as f:
            data = json.load(f)
        relations = [MappingRelation(**d) for d in data]
        print(f"  {len(relations)} relations chargées depuis le cache.")
        return relations

    ref_A_map = {r.id: r for r in ref_A}
    ref_B_map = {r.id: r for r in ref_B}

    relations = asyncio.run(_run_async(candidates, ref_A_map, ref_B_map))

    print(f"\n[Diagnostic étape 4]")
    print(f"  Paires scorées : {len(relations)}")
    for rt in ["equivalence", "A_couvre_B", "B_couvre_A", "partielle", "aucun_lien"]:
        n = sum(1 for r in relations if r.relation_type == rt)
        print(f"  {rt:20s} : {n}")

    _export(relations)
    return relations


# ── Export ────────────────────────────────────────────────────────────────────

def _export(relations: list[MappingRelation]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # JSON
    with open(CACHE_RELATIONS, "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in relations], f, ensure_ascii=False, indent=2)
    print(f"  → {CACHE_RELATIONS}")

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mapping Results"

        headers = [
            "ID Ref A", "Titre Ref A",
            "ID Ref B", "Titre Ref B",
            "Score sémantique", "Coverage A→B", "Coverage B→A",
            "Confiance", "Type de relation",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        # Palette couleur par relation_type
        RELATION_COLORS = {
            "equivalence": "C6EFCE",
            "A_couvre_B":  "FFEB9C",
            "B_couvre_A":  "FFEB9C",
            "partielle":   "FFDDC1",
            "aucun_lien":  "FFC7CE",
        }

        for row_idx, r in enumerate(relations, 2):
            values = [
                r.id_A, r.title_A, r.id_B, r.title_B,
                r.semantic_score, r.coverage_A_to_B, r.coverage_B_to_A,
                r.confidence, r.relation_type,
            ]
            fill_color = RELATION_COLORS.get(r.relation_type, "FFFFFF")
            fill = PatternFill("solid", fgColor=fill_color)
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_idx in (5, 6, 7, 8):
                    cell.number_format = "0.00"
                if col_idx == 9:
                    cell.fill = fill

        # Largeurs de colonnes
        col_widths = [12, 50, 16, 50, 12, 12, 12, 10, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        xlsx_path = OUTPUT_DIR / "mapping_results.xlsx"
        wb.save(str(xlsx_path))
        print(f"  → {xlsx_path}")

    except Exception as e:
        print(f"  [WARN] Export Excel échoué : {e}")
